from openpyxl import *
import pandas as pd

'''wb = pd.read_excel('arquivo.xlsx')'''
'''df.to_excel('teste.xlsx', sheet_name='Teste', na_rep='#N/A', header= True, index=False)'''

wb = load_workbook(filename='arquivo.xls')

sh = wb.active

sh = wb['Controle']

sh.cell(row=2, column=10).value = 'Ano de Nascimento'

sh.cell(row=3, column=10).value = '=SUM(2021-H3)'

print(sh.cell(row=3, column=10).value)

wb.save(filename='arquivocomcalculo.xlsx')